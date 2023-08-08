import openpyxl as op

uzel_name = input("Введите название узла: ")
filename = "Таблица спецификаций 1 3.xlsx"
wb = op.load_workbook(filename, data_only=True)
sheet1 = wb.active

max_rows = sheet1.max_row
max_columns = sheet1.max_column

book = op.Workbook()
sheet = book.active
max_rows1 = sheet.max_row
max_columns1 = sheet.max_column

el = 0
for i in range(4, max_rows + 1):
    sku = sheet1.cell(row=i, column=2).value

    if sku == uzel_name:
        el += 1
        for j in range(1, max_columns):
            sheet.cell(row=el, column=j).value = sheet1[i][j].value

book.save(f"{uzel_name}.xlsx")
book.close()
